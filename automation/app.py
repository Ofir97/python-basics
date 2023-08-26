import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
print(cell.value)

for row in range(2, sheet.max_row + 1):
    corrected_price = sheet.cell(row, 3).value * 0.9
    sheet.cell(row, 4).value = corrected_price

wb.save('transactions2.xlsx')

wb2 = xl.load_workbook('transactions2.xlsx')
sheet = wb2['Sheet1']

for row in range(2, sheet.max_row + 1):
    print(sheet.cell(row, 4).value)
